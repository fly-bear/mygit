from PyQt5 import QtWidgets, QtCore, QtGui
import sys,os
from biao import Ui_MainWindow
from PyQt5.QtWidgets import QFileDialog,QMessageBox
from PyQt5.QtCore import QCoreApplication
import time
import upload
from openpyxl import load_workbook

class mywindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(mywindow, self).__init__()
        self.setupUi(self)
        self.comboBox.addItem("西院")
        self.comboBox.addItem("东院")
        self.comboBox_2.addItem("是")
        self.comboBox_2.addItem("否")
        self.comboBox_3.addItem("是")
        self.comboBox_3.addItem("否")
        # self.dateTimeEdit.setCalendarPopup(True)
        # self.dateTimeEdit_2.setCalendarPopup(True)
        # self.dateEdit.setCalendarPopup(True)
        now_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        self.dateTimeEdit.setDateTime(QtCore.QDateTime.fromString(now_time, 'yyyy-MM-dd hh:mm:ss'))
        self.dateTimeEdit_2.setDateTime(QtCore.QDateTime.fromString(now_time, 'yyyy-MM-dd hh:mm:ss'))
        now_day = time.strftime("%Y-%m-%d", time.localtime())
        self.dateEdit.setDate(QtCore.QDate.fromString(now_day, 'yyyy-MM-dd'))
        self.lineEdit_4.setText("狮城412")

    def cb1(self):
        self.lineEdit_3.setText(self.comboBox.currentText())

    def cb2(self):
        self.lineEdit_5.setText(self.comboBox_2.currentText())

    def cb3(self):
        self.lineEdit_12.setText(self.comboBox_3.currentText())

    def get_message(self):
        message = ['自动化学院', '2015级', '自动化1502班']
        message.append(self.lineEdit.text())
        for i in range(2, 8):
            nowt = 'self.lineEdit_' + str(i) + '.text()'
            s = eval(nowt)
            message.append(s)
            minute1=self.dateTimeEdit.dateTime().toPyDateTime().minute
            if minute1<10:
                minute1='0'+str(minute1)
            else:
                minute1=str(minute1)
            minute2 = self.dateTimeEdit_2.dateTime().toPyDateTime().minute
            if minute2 < 10:
                minute2 = '0' + str(minute2)
            else:
                minute2 = str(minute2)
        message.append(str(self.dateTimeEdit.dateTime().toPyDateTime().year)+'/'+
                       str(self.dateTimeEdit.dateTime().toPyDateTime().month)+'/'+
                       str(self.dateTimeEdit.dateTime().toPyDateTime().day))
        message.append(str(self.dateTimeEdit.dateTime().toPyDateTime().hour) + ':' +
                       minute1)
        message.append(str(self.dateTimeEdit_2.dateTime().toPyDateTime().year) + '/' +
                       str(self.dateTimeEdit_2.dateTime().toPyDateTime().month) + '/' +
                       str(self.dateTimeEdit_2.dateTime().toPyDateTime().day))
        message.append(str(self.dateTimeEdit_2.dateTime().toPyDateTime().hour) + ':' +
                       minute2)
        for i in range(8, 13):
            nowt = 'self.lineEdit_' + str(i) + '.text()'
            s = eval(nowt)
            message.append(s)
        message.append(self.dateEdit.date().toPyDate().year)
        message.append(self.dateEdit.date().toPyDate().month)
        message.append(self.dateEdit.date().toPyDate().day)
        message=message[:15]+[self.lineEdit_13.text()]+message[15:]
        return message

    def handle(self):
        message=self.get_message()
        print(message)
        filename=str(message[-3])+str(message[-2])+str(message[-1])+'.xlsx'
        file=upload.get_file(filename)
        with open(filename,'wb') as f:
            f.write(file)
        wb = load_workbook(filename)
        sheet = wb.get_sheet_by_name('Sheet1')
        # rows=sheet.max_row
        use=message[:-3]
        sheet.append(use)
        wb.save(filename)
        upload.upload_file(filename)
        apply = QMessageBox.question(self,
                                     '完成',
                                     '上传完成，是否退出？',
                                     QMessageBox.Yes | QMessageBox.No)
        if apply==QMessageBox.Yes:
            QCoreApplication.quit()
        else:
            pass




if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = mywindow()
    window.show()
    sys.exit(app.exec_())
